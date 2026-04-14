import os
import sys
import time
import logging
import requests
from dotenv import load_dotenv
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')
from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# ==========================================
# CONFIGURAÇÃO DE SEGURANÇA E AMBIENTE
# ==========================================
# 1. Força o caminho ABSOLUTO (resolve o erro do subprocess)
pasta_raiz = os.path.dirname(os.path.abspath(__file__))
caminho_env = os.path.join(pasta_raiz, '.env')

# 2. Carrega os segredos forçando a leitura do arquivo correto
load_dotenv(caminho_env, override=True)

# Resgata as chaves
TOKEN_TELEGRAM = os.getenv("TOKEN_TELEGRAM")
CHAT_ID_ARLINDO = os.getenv("CHAT_ID_ARLINDO")

# Lê a meta de pontos e converte para número
META_PONTOS = int(os.getenv("META_PONTOS", "120000"))

# Lê o caminho do perfil do Chrome
CAMINHO_CHROME = os.getenv("CAMINHO_PERFIL_CHROME")

# Validação de segurança
if not TOKEN_TELEGRAM or not CHAT_ID_ARLINDO:
    logging.error(f"❌ Erro Crítico: O arquivo .env não foi lido corretamente no caminho: {caminho_env}")
    raise ValueError("Credenciais do Telegram em falta. Verifique o seu .env.")

# ==========================================
# FUNÇÃO DE ALERTA DO ROBÔ
# ==========================================
def enviar_telegram(mensagem):
    try:
        url = f"https://api.telegram.org/bot{TOKEN_TELEGRAM}/sendMessage"
        payload = {
            "chat_id": CHAT_ID_ARLINDO,
            "text": mensagem,
            "parse_mode": "Markdown"
        }
        
        response = requests.post(url, json=payload)

        if response.status_code == 200:
            logging.info("[TELEGRAM] Alerta enviado com sucesso!")
        else:
            logging.error(f"[TELEGRAM] Erro na API: {response.text}")
            
    except Exception as e:
        logging.error(f"[TELEGRAM] Erro de conexão: {e}")

def salvar_historico_excel(origem, destino, data_voo_formatada, menor_preco):
    arquivo_excel = os.getenv("CAMINHO_PLANILHA")
    agora = datetime.now()
    
    try:
        if not os.path.exists(arquivo_excel):
            wb = Workbook()
            ws = wb.active
            ws.title = "Historico_Precos"
            ws.append(["Data/Hora da Busca", "Origem", "Destino", "Data do Voo", "Preço (Pontos)"])
        else:
            wb = load_workbook(arquivo_excel)
            ws = wb["Historico_Precos"]
        
        primeira_linha_vazia = 2
        for row in range(2, ws.max_row + 100): 
            valor_celula = ws.cell(row=row, column=1).value
            if valor_celula is None or str(valor_celula).strip() == "":
                primeira_linha_vazia = row
                break
        
        celula_data = ws.cell(row=primeira_linha_vazia, column=1, value=agora)
        celula_data.number_format = 'dd/mm/yyyy hh:mm:ss'
        
        ws.cell(row=primeira_linha_vazia, column=2, value=origem)
        ws.cell(row=primeira_linha_vazia, column=3, value=destino)
        ws.cell(row=primeira_linha_vazia, column=4, value=data_voo_formatada)
        
        celula_preco = ws.cell(row=primeira_linha_vazia, column=5, value=menor_preco)
        celula_preco.number_format = '#,##0' 
        
        wb.save(arquivo_excel)
        logging.info(f"[PLANILHA] Histórico salvo na linha {primeira_linha_vazia}: {data_voo_formatada} - {menor_preco} pontos.")
        
    except PermissionError:
        logging.error("🚨 [ERRO] O arquivo Excel está ABERTO no seu computador! Feche o arquivo.")
    except Exception as e:
        logging.error(f"🚨 [ERRO] Falha ao salvar no Excel: {e}")

# ==========================================
# CONFIGURAÇÃO DO LOG
# ==========================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("log_azul_bot.txt", mode='w', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

def buscar_passagem_azul(page, origem, destino, data_ida):
    try:
        page.goto("https://www.voeazul.com.br")

        try:
            page.locator("#onetrust-accept-btn-handler").click(timeout=3000)
        except: pass

        time.sleep(3) 
        page.keyboard.press("Escape")
        time.sleep(0.5)
        page.keyboard.press("Escape")
        time.sleep(1)

        campo_origem = page.locator("input[aria-label='Origem'][role='combobox']").first
        campo_origem.evaluate("node => node.focus()")
        campo_origem.evaluate("node => node.click()")
        time.sleep(0.5)
        page.keyboard.press("Control+A")
        page.keyboard.press("Backspace")
        campo_origem.press_sequentially(origem, delay=200) 
        time.sleep(2) 
        page.keyboard.press("ArrowDown")
        time.sleep(0.5)
        page.keyboard.press("Enter")

        campo_destino = page.locator("input[aria-label='Destino'][role='combobox']").first
        campo_destino.evaluate("node => node.focus()")
        campo_destino.evaluate("node => node.click()")
        time.sleep(0.5)
        page.keyboard.press("Control+A")
        page.keyboard.press("Backspace")
        campo_destino.press_sequentially(destino, delay=200)
        time.sleep(2)
        for _ in range(3):
            page.keyboard.press("ArrowDown")
            time.sleep(0.3) 
        time.sleep(0.5)
        page.keyboard.press("Enter")

        campo_data = page.locator("input[aria-label='Data de ida']").first
        campo_data.evaluate("node => node.focus()")
        campo_data.evaluate("node => node.click()")
        time.sleep(0.5)
        page.keyboard.press("Control+A")
        page.keyboard.press("Backspace")
        campo_data.press_sequentially(data_ida, delay=150)
        time.sleep(1)
        page.keyboard.press("Enter")
        page.keyboard.press("Escape")
        time.sleep(1)

        try:
            page.locator("input[data-cy='checkbox-points']").evaluate("node => node.click()")
        except:
            try: page.locator("text='Usar pontos Azul'").evaluate("node => node.click()")
            except: pass
        time.sleep(1)

        try:
            page.locator("button:has-text('Buscar passagens')").evaluate("node => node.click()")
        except:
            page.keyboard.press("Enter") 

        logging.info("Aguardando o carregamento dos voos...")
        seletor_preco = "[data-test-id*='fare-price-with-points']"
        
        page.wait_for_selector(seletor_preco, timeout=30000)
        time.sleep(2) 
        
        elementos_preco = page.locator(seletor_preco).all()
        lista_de_precos = []
        
        for elemento in elementos_preco:
            texto_limpo = elemento.inner_text().replace("pontos", "").replace(".", "").strip()
            try:
                lista_de_precos.append(int(texto_limpo))
            except ValueError:
                continue 
        
        if len(lista_de_precos) > 0:
            menor_preco = min(lista_de_precos)
            data_formatada = f"{data_ida[:2]}/{data_ida[2:4]}/{data_ida[4:]}"
            preco_formatado = f"{menor_preco:,}".replace(",", ".")
            meta_formatada = f"{META_PONTOS:,}".replace(",", ".")
            
            logging.info(f" -> O MENOR PREÇO DO DIA {data_formatada} É: {preco_formatado} PONTOS! <- ")

            salvar_historico_excel(origem, destino, data_formatada, menor_preco)

            if menor_preco <= META_PONTOS:
                mensagem_alerta = (
                    f"🚨 *OPORTUNIDADE AZUL ENCONTRADA* 🚨\n\n"
                    f"📍 *Origem:* {origem}\n"
                    f"🏁 *Destino:* {destino}\n"
                    f"📅 *Data:* {data_formatada}\n"
                    f"💰 *Melhor Preço:* {preco_formatado} pontos\n\n"
                    f"⚠️ O valor está abaixo da meta de {meta_formatada} pontos!"
                )
                enviar_telegram(mensagem_alerta)
            return True 
        else:
            raise ValueError("Nenhum valor extraído da tela.")

    except Exception as erro:
        logging.error(f"Erro durante a extração: {erro}")
        raise 

if __name__ == "__main__":
    SIGLA_ORIGEM = "VCP"
    SIGLA_DESTINO = "LIS"
    MAX_TENTATIVAS = 3 

    # ==========================================
    # LEITURA E VALIDAÇÃO DO ARQUIVO DE DATAS
    # ==========================================
    DATAS_VIAGEM = []
    
    # 3. Força o caminho absoluto também para as datas
    caminho_datas = os.path.join(pasta_raiz, "datas_viagem.txt")

    texto_manual = (
        "=== MANUAL DE DATAS PARA O ROBÔ ===\n"
        "# O robô só fará a leitura de linhas que contenham o caractere \":\"\n"
        "# Digite a data com 8 dígitos, sem barras.\n\n"
        "Data alvo: 04092026\n"
        "Data alvo: 05092026\n"
        "Data alvo: 06092026\n"
    )

    if not os.path.exists(caminho_datas):
        logging.info("⚙️ Arquivo 'datas_viagem.txt' não encontrado. Criando arquivo com o manual padrão...")
        with open(caminho_datas, "w", encoding="utf-8") as f:
            f.write(texto_manual)

    with open(caminho_datas, "r", encoding="utf-8") as f:
        linhas = f.readlines()
        
        for linha in linhas:
            if not linha.strip() or linha.startswith("#") or linha.startswith("="):
                continue
                
            if ":" in linha:
                data_extraida = linha.split(":")[1].strip()
                if len(data_extraida) == 8 and data_extraida.isdigit():
                    DATAS_VIAGEM.append(data_extraida)
                else:
                    logging.warning(f"⚠️ Formato inválido ignorado na leitura: '{data_extraida}'")

    if len(DATAS_VIAGEM) == 0:
        logging.warning("⚠️ Nenhuma data válida encontrada no arquivo. Restaurando o manual de segurança...")
        DATAS_VIAGEM = ["04092026", "05092026", "06092026"]
        with open(caminho_datas, "w", encoding="utf-8") as f:
            f.write(texto_manual)

    logging.info(f"✅ Robô inicializado. Datas validadas: {DATAS_VIAGEM}")

    # ==========================================
    # SESSÃO ÚNICA: Abre o navegador apenas uma vez
    # ==========================================
    with sync_playwright() as p:
        browser = p.chromium.launch_persistent_context(
            user_data_dir=CAMINHO_CHROME, # <-- A variável entra aqui sem aspas
            channel="chrome",
            headless=False,
            args=["--start-maximized"]
        )
        
        # A primeira aba que o navegador abre
        page = browser.pages[0]

        for data in DATAS_VIAGEM:
            logging.info(f"\n==========================================")
            logging.info(f" INICIANDO FLUXO PARA A DATA: {data}")
            logging.info(f"==========================================")
            
            sucesso_na_data = False
            
            for tentativa in range(1, MAX_TENTATIVAS + 1):
                logging.info(f"Tentativa {tentativa} de {MAX_TENTATIVAS}...")
                try:
                    buscar_passagem_azul(page, SIGLA_ORIGEM, SIGLA_DESTINO, data)
                    sucesso_na_data = True
                    break 
                except Exception as e:
                    logging.warning(f"Falha na tentativa {tentativa} para a data {data}. Erro: {e}")
                    if tentativa < MAX_TENTATIVAS:
                        logging.info("Aguardando 5 segundos antes de tentar novamente...")
                        time.sleep(5)
            
            if not sucesso_na_data:
                msg_erro = f"⚠️ *Alerta do Robô*\nNão consegui extrair os preços para o dia {data[:2]}/{data[2:4]} após {MAX_TENTATIVAS} tentativas."
                enviar_telegram(msg_erro)
                logging.error(f"Desistindo da data {data} após {MAX_TENTATIVAS} tentativas.")

        logging.info("\n==========================================")
        logging.info("Fechando o navegador...")
        browser.close()

    logging.info("TODAS AS DATAS PROCESSADAS.")
